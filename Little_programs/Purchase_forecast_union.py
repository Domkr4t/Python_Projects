import pandas as pd
import os
import math
import platform
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import csv
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk


sales_file_path = None
residuals_file_path = None
report_file_path = None
selected_store = None


all_beer = ["Жигулёвское", "Светлое НФ", "Тёмное", "Пилснер Ф", "Пилснер НФ", "Светлое Ф", "Пшеничное", "Вишнёвый крик", "Хеллес", "Крепкое", "Лёгкое", "IPA", "Грейпфрутовый эль", "APA", "Квас Воронеж"]

merka_snacks = ["Гренки Бородинские Чеснок", "Гренки по-деревенски чеснок", "Гренки тайский соус", "Гренки багет Мексиканский микс", "Гренки красная икра", "Гренки томат-зелень", "Гренки чеснок", "Гренки сыр", "Ломтики курицы", "Ломтики говядины", "Чипсы мясные свинина", "Чипсы мясные курица", "Курица Халапеньо", "Чипсы мясные индейка гриль", "Фисташки", "Миндаль жареный соленый", "Ореховый микс", "Японские снэки", "Арахис семга-сыр", "Арахис в глазури сметана-зелень", "Арахис в глазури сыр", "Арахис шашлык", "Арахис семга - сыр", "Арахис соль", "Арахис сыр-чеснок", "Арахис в глазури васаби", "Сыр Косичка", "Сыр Охотник", 'Сыр "Бочонок"', "Сыр Джил", "Семечки с солью 130г", "Уши свиные в ассортименте 90г", "Чипсы Мистер Потато оригинальные 40г", "Чипсы Мистер Потато сметана/лук 40г", "Чипсы Мистер Потато барбекю 40г", "Чипсы Мистер Потато острые 40г", "Лещ", "Камбала с икрой", "Камбала Ёрш", "Камбала без икры", "Пелядь", "Чехонь", "Плотва", "Синец", "Вобла", "Мойва вяленая", "Сырок", "Щука", "Вобла Астраханская", "Рыбец", "Тарань", "Палочки кеты", "Мясо краба", "Желтый полосатик", "Икра минтая", "Осьминог", "Мясо краба по-шанхайски", "Кольца кальмара", "Палочки горбуши", "Кольца кальмара по-шанхайски", "Стружка кальмара", "Мидии", "Ассорти рыбное", "Стружка кальмара по-шанхайски", "Икра воблы", "Хот-тейс", "Камбалка сушеная", "Колбаски мясные со вкусом чили", "Колбаски мясные с чесноком", "Соломка форели", "Таранка с перцем", "Щупальцы кальмара", "Вомер х/к", "Жерех х/к", "Теша горбуши х/к", "Лещ х/к]"]

kaspi_snacks = ["Креветка сушеная с чесноком и укропом 40г", "Креветка сушеная с солью  40г", "Гренки Волнистые с чесноком 75г", "Гренки Барные с томатом,чесноком и зеленью 70г", "Корюшка без икры", "Корюшка с икрой", "Иваси тушка х/к", "Киперс х/к", "Гренки Живые с чесноком", "Соломка семги", "Соломка воблы", "Бобы жареные чили", "Бобы жареные мексико", "Бобы жареные чеснок", "Бобы жареные соль", "Гренки Волнистые с паприкой 75г"]

piece_snacks = ["Уши свиные в ассортименте 90г", "Чипсы Мистер Потато барбекю 40г", "Чипсы Мистер Потато оригинальные 40г", "Чипсы Мистер Потато острые 40г", "Чипсы Мистер Потато сметана/лук 40г", "Гренки Волнистые с чесноком 75г", "Креветка сушеная с солью  40г", "Креветка сушеная с чесноком и укропом 40г", "Гренки Барные с томатом,чесноком и зеленью 70г", "Семечки с солью 130г", "Гренки Волнистые с паприкой 75г"]

sigma_snacks = ["Сиг г/к"]

banki = ["Банка 1л", "Банка 2л", "Банка 3л", "Крышка"]

static_nomenclature = ["Пакет плотный -", "Пакет майка -", "Перчатки одноразовые -", "Пакет фасовочный -", "Контейнер черный новый -","Контейнер 250 мл. -", "Контейнер 500 мл. -", "Контейнер 1000 мл. -", "Мусорные пакеты большие -", "Мусорные пакеты маленькие на завязке  -", "Лента узкая -", "Стакан 0.5 -", "Ручка для банки -"]

kaspi_green = ["Корюшка без икры", "Корюшка с икрой", "Иваси тушка х/к", "Киперс х/к", "Гренки Живые с чесноком", "Соломка воблы"]

kaspi_yellow = ["Соломка семги", "Креветка сушеная с чесноком и укропом 40г", "Креветка сушеная с солью  40г", "Гренки Волнистые с чесноком 75г", "Гренки Барные с томатом,чесноком и зеленью 70г"]

merka_green = ["Гренки по-деревенски чеснок", "Гренки тайский соус", "Ломтики курицы", "Ломтики говядины", "Чипсы мясные свинина", "Чипсы мясные курица", "Чипсы мясные индейка гриль", "Фисташки", "Миндаль жареный соленый", "Ореховый микс", "Японские снэки", "Арахис семга-сыр", "Арахис в глазури сметана-зелень", "Арахис в глазури сыр", "Арахис шашлык", "Арахис соль", "Сыр Косичка", "Сыр Охотник", 'Сыр "Бочонок"', "Сыр Джил", "Семечки с солью 130г", "Уши свиные в ассортименте 90г", "Чипсы Мистер Потато оригинальные 40г", "Чипсы Мистер Потато сметана/лук 40г", "Камбала с икрой", "Камбала Ёрш", "Камбала без икры", "Пелядь", "Чехонь", "Плотва", "Синец", "Вобла", "Мойва вяленая", "Лещ", "Палочки кеты", "Мясо краба", "Желтый полосатик", "Икра минтая", "Осьминог", "Мясо краба по-шанхайски", "Кольца кальмара", "Палочки горбуши", "Кольца кальмара по-шанхайски", "Стружка кальмара", "Мидии", "Ассорти рыбное", "Стружка кальмара по-шанхайски", "Икра воблы", "Хот-тейс", "Камбалка деликатесная", "Колбаски мясные со вкусом чили", "Колбаски мясные с чесноком", "Вомер х/к", "Жерех х/к", 'Теша горбуши х/к', "Камбалка сушеная"]

merka_yellow = ["Гренки багет Мексиканский микс", "Гренки красная икра", "Гренки томат-зелень", "Гренки чеснок", "Гренки сыр", "Курица Халапеньо", "Арахис сыр-чеснок", "Арахис в глазури васаби", "Чипсы Мистер Потато барбекю 40г", "Чипсы Мистер Потато острые 40г", "Щука", "Вобла Астраханская", "Рыбец", "Тарань", "Сырок", "Соломка форели", "Таранка с перцем", "Щупальцы кальмара", "Лещ х/к"]


fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#Функция округления
def custom_ceil(number):
    if number - int(number) >= 0.15:
        return math.ceil(number)
    else:
        return math.floor(number)


# Функция для обработки пива
def handle_beer(nomenclature, stock, total_quantity):
    nomenclature_50 = ["Пилснер НФ", "Пилснер Ф", "Жигулёвское", "Квас"]

    if nomenclature in nomenclature_50:
        forecast = math.floor((stock - total_quantity) / 50)
    elif nomenclature == "Вишнёвый крик":
        forecast = math.floor((stock - total_quantity) / 20)
    elif nomenclature == "Вице Канцлер б/а  0,45":
        forecast = math.ceil(total_quantity / 12)
    else:
        forecast = math.floor((stock - total_quantity) / 30)

    if (forecast >= 0):
        forecast = 0
        remaining_liters = stock - total_quantity
    else:
        remaining_liters = 0

    if remaining_liters <= 25:
        if forecast <= 0:
            forecast = forecast - 1
        else:
            forecast = forecast + 1

    return {"Номенклатура": nomenclature, "Остаток": stock, "Прогноз": total_quantity, "Заказ кег": forecast, "Остаток литров": remaining_liters}


# Функция для обработки закусок к пиву
def handle_snacks(nomenclature, stock, total_quantity):
    forecasted_balance = stock - total_quantity

    if nomenclature in piece_snacks:
        if nomenclature == "Гренки Волнистые с чесноком 75г" or nomenclature == "Гренки Волнистые с паприкой 75г":
            if forecasted_balance >= 5:
                forecast = 0
            elif forecasted_balance < 5:
                forecast = 14
        else:
            if forecasted_balance >= 5:
                forecast = 0
            elif forecasted_balance < 5:
                forecast = 5
    elif nomenclature not in piece_snacks:
        if total_quantity <= 0.100 and forecasted_balance >= 0.300:
            forecast = 0
        elif forecasted_balance < 0.600:
            forecast = 1 - abs(forecasted_balance)
        elif forecasted_balance < 0:
            forecast = abs(forecasted_balance)
        elif forecasted_balance >= 0.600:
            forecast = 0


    return {"Номенклатура": nomenclature, "Остаток": stock, "Прогноз": total_quantity, "Прогнозируемый остаток": forecasted_balance, "Заказ": abs(forecast)}


# Функция для обработки прочего
def handle_other(nomenclature, stock, total_quantity):
    forecasted_balance = stock - total_quantity
    if nomenclature == "Банка 1л":
        forecast = math.ceil((total_quantity + 24)/12)
    elif nomenclature == "Банка 2л" or nomenclature == "Банка 3л":
        forecast = math.ceil((total_quantity + 12)/6)
    elif nomenclature == "Крышка":
        if forecasted_balance < 200:
            forecast = 1
        else:
            forecast = 0
    else:
        return None

    return {"Номенклатура": nomenclature, "Остаток": stock, "Прогноз": total_quantity, "Прогнозируемый остаток": forecasted_balance, "Заказ": abs(forecast)}


def calc_end_date_kaspi():
    day_of_week = datetime.now().weekday()
    start_date_order = datetime.now() - timedelta(days=7)

    if day_of_week == 2:
        if selected_store == "12_Фрязино_Мира8":
            return ((start_date_order + timedelta(days=6))).replace(hour=23, minute=59, second=59, microsecond=59)
        else:
            return datetime.max
    elif day_of_week == 3:
        return ((start_date_order + timedelta(days=5))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 6:
        return ((start_date_order + timedelta(days=7))).replace(hour=23, minute=59, second=59, microsecond=59)
    else:
        return datetime.max


def calc_end_date_merka():
    day_of_week = datetime.now().weekday()
    start_date_order = datetime.now() - timedelta(days=7)

    if day_of_week == 0:
        if selected_store == "10_Коломна_Советская5" or selected_store == "17_Коломна_Кирова38":
            return ((start_date_order + timedelta(days=8))).replace(hour=23, minute=59, second=59, microsecond=59)
        else:
            return ((start_date_order + timedelta(days=3))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 1:
        return ((start_date_order + timedelta(days=3))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 2:
        return ((start_date_order + timedelta(days=5))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 3:
        return ((start_date_order + timedelta(days=6))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 6:
        if selected_store == "14_Егорьевск_Советская191":
            return ((start_date_order + timedelta(days=7))).replace(hour=23, minute=59, second=59, microsecond=59)
        else:
            return ((start_date_order + timedelta(days=4))).replace(hour=23, minute=59, second=59, microsecond=59)
    else:
        return datetime.max


def calc_end_date_beer():
    day_of_week = datetime.now().weekday()
    start_date_order = datetime.now() - timedelta(days=7)

    if day_of_week == 0:
        return ((start_date_order + timedelta(days=3))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 1:
        return ((start_date_order + timedelta(days=3))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 2:
        if selected_store == "9_Балашиха_Советский6/17":
            return ((start_date_order + timedelta(days=5))).replace(hour=23, minute=59, second=59, microsecond=59)
        else:
            return ((start_date_order + timedelta(days=3))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 3:
        return ((start_date_order + timedelta(days=4))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 4:
        if selected_store == "7_Балашиха_Свердлова25" or selected_store == "18_Железнодорожный":
            return ((start_date_order + timedelta(days=3))).replace(hour=23, minute=59, second=59, microsecond=59)
        else:
            return ((start_date_order + timedelta(days=2))).replace(hour=23, minute=59, second=59, microsecond=59)
    elif day_of_week == 6:
        return ((start_date_order + timedelta(days=3))).replace(hour=23, minute=59, second=59, microsecond=59)
    else:
        return datetime.max


def calc_end_date_other():
    day_of_week = datetime.now().weekday()
    start_date_order = datetime.now() - timedelta(days=7)

    if day_of_week == 6:
        return ((start_date_order + timedelta(days=7))).replace(hour=23, minute=59, second=59, microsecond=59)
    else:
        return datetime.max


def text_for_shop():
    if selected_store == "6_Люберцы_3-е Почтовое отделение74":
        return ["Добрый день.", "Заказ ИП Аганина,", "улица 3-е Почтовое Отделение 74,  Люберцы,"]
    elif selected_store == "16_Долгопрудный_Лихачевский68":
        return ["Добрый день.", "Заказ ИП Аганина,", "Лихачёвский проспект 68, Долгопрудный,"]
    elif selected_store == "5_Балашиха_Фадеева3":
        return ["Добрый день.", "Заказ ИП Петрова, ", "Ул.Фадеева, 3А, Балашиха,"]
    elif selected_store == "7_Балашиха_Свердлова25":
        return ["Добрый день.", "Заказ ИП Аганина, ", "Ул.Свердлова25а, Балашиха,"]
    elif selected_store == "4_Электросталь_Ленина15":
        return ["Добрый день.", "Заказ ИП Петрова,", "проспект Ленина, 15, Электросталь, "]
    elif selected_store == "3_Электросталь_Ялагина11":
        return ["Добрый день.", "Заказ ИП Петрова,", "Ул.Ялагина, 11, Электросталь,"]
    elif selected_store == "9_Балашиха_Советский6/17":
        return ["Добрый день.", "Заказ ИП Петрова,", "Ул.Советская6/17, Балашиха,"]
    elif selected_store == "14_Егорьевск_Советская191":
        return ["Добрый день.", "Заказ ИП Аганина,", "Советская улица, 191, Егорьевск,"]
    elif selected_store == "10_Коломна_Советская5":
        return ["Добрый день.", "Заказ ИП Петрова ", "Ул.Советская площадь, 5А, Коломна,"]
    elif selected_store == "12_Фрязино_Мира8":
        return ["Добрый день.", "Заказ ИП Петрова, ", "Пр-кт Мира 8, Фрязино,"]
    elif selected_store == "1_Дрезна_Южная19а":
        return ["Добрый день.", "Заказ ИП Петрова,", "Южная улица, 19А, Дрезна,"]
    elif selected_store == "3_Электросталь_Ялагина11":
        return ["Добрый день.", "Заказ ИП Петрова,", "Ул.Ялагина, 11, Электросталь,"]
    elif selected_store == "2_Электросталь_Победы1/2":
        return ["Добрый день.", "Заказ ИП Петрова,", "улица Победы, 1к2, Электросталь,"]
    elif selected_store == "17_Коломна_Кирова38":
        return ["Добрый день.", "Заказ ИП Петрова,", "Г. Коломна Проспект Кирова 38а"]
    elif selected_store == "18_Железнодорожный":
        return ["Добрый день.", "Заказ ИП Петрова,", "Г. Балашиха, мкр. Железнодорожный, ул. Маяковского, дом 12 стр 2"]
    else:
        return "Хз чо за магаз"


def fill_excel_kaspi(ws):
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            for word in kaspi_green:
                if str(cell.value).lower() == word.lower():
                    cell.fill = fill_green
                    break
            for word in kaspi_yellow:
                if str(cell.value).lower() == word.lower():
                    cell.fill = fill_yellow
                    break


def fill_excel_merka(ws):
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            for word in merka_green:
                if str(cell.value).lower() == word.lower():
                    cell.fill = fill_green
                    break
            for word in merka_yellow:
                if str(cell.value).lower() == word.lower():
                    cell.fill = fill_yellow
                    break

# Загрузка файла с продажами
def load_sales_file():
    global sales_file_path
    sales_file_path = filedialog.askopenfilename(initialdir="c:/User/Desktop", title="Загрузка файла с продажами")
    print(f"Выбран файл с продажами: {sales_file_path}")
    if sales_file_path:
        message_label_sales.config(text=f"{os.path.basename(sales_file_path)}", foreground="green")

        # Читаем уникальные значения магазинов из файла
        sales_data = pd.read_excel(sales_file_path, skiprows=4, usecols=[0])

        sales_data_column_names = ["Магазин"]
        sales_data.columns = sales_data_column_names

        stores = sales_data["Магазин"].unique()
        stores = [store.strip() for store in stores]

        # Обновляем выпадающий список магазинов
        store_combobox["values"] = stores[0::]
        store_combobox.set(stores[0])  # Устанавливаем первый магазин по умолчанию

        message_error.config(text=" ")
    else:
        message_label_sales.config(text="Файл не выбран", foreground="red")
        store_combobox["values"] = ""
        store_combobox.set("")
        sales_file_path = None


# Загрузка файла с остатками
def load_residuals_file():
    global residuals_file_path
    residuals_file_path = filedialog.askopenfilename(initialdir="c:/User/Desktop", title="Загрузка файла с остатками")
    print(f"Выбран файл с остатками: {residuals_file_path}")
    if residuals_file_path:
        message_label_residuals.config(text=f"{os.path.basename(residuals_file_path)}", foreground="green")
        message_error.config(text=" ")
    else:
        message_label_residuals.config(text="Файл не выбран", foreground="red")
        residuals_file_path = None


# Открытие файла
def open_file(file_path):
    # Открывает файл с помощью системной команды.
    if platform.system() == "Windows":
        os.startfile(file_path)
    elif platform.system() == "Darwin":
        os.system(f"open {file_path}")
    else:
        os.system(f"xdg-open {file_path}")


# Сохранение файла
def save_file():
    global report_file_path
    file_name_start_date = start_date.get()
    file_name_end_date = end_date.get()

    # Проверка на то что дата пустая строка
    if not file_name_start_date and not file_name_end_date:
        file_default_name = f"Закупки_по_{selected_store}"
    elif not file_name_end_date:
        file_default_name = f"Закупки_по_{selected_store}_c_{file_name_start_date}"
    elif not file_name_start_date:
        file_default_name = f"Закупки_по_{selected_store}_по_{file_name_end_date}"
    else:
        file_default_name = f"Закупки_по_{selected_store}_c_{file_name_start_date}_по_{file_name_end_date}"

    file_default_name = file_default_name.replace("/", "_").replace("\\", "_")

    # Открываем диалог выбора файла для сохранения
    report_file_path = filedialog.asksaveasfilename(initialfile=file_default_name, defaultextension=".xlsx", title="Сохранить отчет", filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])


# Создание отчета
def generate_report():
    global selected_store
    global start_date
    global end_date

    # Проверка на заполнение файлов
    if sales_file_path is None and residuals_file_path is None:
        message_error.config(text="Файлы не загружены")
        return
    elif sales_file_path is None:
        message_error.config(text="Продажи не загружены")
        return
    elif residuals_file_path is None:
        message_error.config(text="Остатки не загружены")
        return

    # Чтение данных из первого файла (с продажами)
    sales = pd.read_excel(f"{sales_file_path}", skiprows=4, usecols=[0, 3, 4, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18])
    sales_column_names = ["Магазин", "Дата и время", "Id чека", "Диск карт", "Владелец карты", "Номер телефона", "Категория товара", "Группа товара", "Номенклатура", "Сумма продаж", "Количество товара", "Остаток на складе", "Сумма скидки", "Себестоимость продаж", "Валовая прибыль"]
    sales.columns = sales_column_names
    sales['Дата и время'] = pd.to_datetime(sales['Дата и время'], dayfirst=True)

    # Чтение данных из второго файла (с остатками)
    residuals = pd.read_excel(f"{residuals_file_path}", skiprows=2, usecols=[0, 2]) # пропуск строки с назва
    residuals_column_names = ["Номенклатура", "Остаток"]
    residuals.columns = residuals_column_names

    # Создание "шапок" для таблиц
    beer_results = pd.DataFrame(columns=["Номенклатура", "Остаток", "Прогноз", "Заказ кег", "Остаток литров"])
    beer_second_table = pd.DataFrame(columns=["Номенклатура(Воронеж)", "Заказ кег"])
    beer_second_table_our_name = pd.DataFrame(columns=["Номенклатура(Воронеж)", "Заказ кег"])
    snacks_results_kaspi = pd.DataFrame(columns=["Номенклатура", "Остаток", "Прогноз", "Прогнозируемый остаток", "Заказ"])
    snacks_results_merka = pd.DataFrame(columns=["Номенклатура", "Остаток", "Прогноз", "Прогнозируемый остаток", "Заказ"])
    # snacks_results_sigma = pd.DataFrame(columns=["Номенклатура", "Остаток", "Прогноз", "Прогнозируемый остаток", "Заказ"])
    snacks_second_table = pd.DataFrame(columns=["Номенклатура(Мерка)", "Заказ"])
    snacks_second_table_kaspi = pd.DataFrame(columns=["Номенклатура(Каспи)", "Заказ"])
    # snacks_second_table_sigma = pd.DataFrame(columns=["Номенклатура(Сиг)", "Заказ"])
    other_results = pd.DataFrame(columns=["Номенклатура", "Остаток", "Прогноз", "Прогнозируемый остаток", "Заказ"])
    other_second_table = pd.DataFrame(columns=["Номенклатура", "Заказ"])
    text_shop = pd.DataFrame()
    start_date_in_doc = pd.DataFrame()
    end_date_in_doc_kaspi = pd.DataFrame()
    end_date_in_doc_merka = pd.DataFrame()
    end_date_in_doc_beer = pd.DataFrame()
    end_date_in_doc_other = pd.DataFrame()

    selected_store = store_combobox.get()

    start_date_str = start_date.get() + " 00:00:00" if start_date.get() else "01.01.1900 00:00:00"
    end_date_str = end_date.get() + " 23:59:59" if end_date.get() else "31.12.9999 23:59:59"
    try:
        start_date_dt = datetime.strptime(start_date_str, "%d.%m.%Y %H:%M:%S")
    except ValueError:
        message_error.config(text="Введен неправильный формат начальной даты")

    try:
        end_date_dt = datetime.strptime(end_date_str, "%d.%m.%Y %H:%M:%S")
    except ValueError:
        message_error.config(text="Введен неправильный формат конечной даты")

    start_date_str = start_date_dt.strftime("%d.%m.%Y %H:%M:%S").format()
    end_date_str = end_date_dt.strftime("%d.%m.%Y %H:%M:%S").format()

    start_date_order = (datetime.now() - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)

    end_date_kaspi = calc_end_date_kaspi()
    end_date_merka = calc_end_date_merka()
    end_date_beer = calc_end_date_beer()
    end_date_other = calc_end_date_other()

    all_nomenclatures_from_sales = set(sales["Номенклатура"].unique())
    residuals_nomenclatures = set(residuals["Номенклатура"])
    missing_in_residuals = all_nomenclatures_from_sales - residuals_nomenclatures

    for missing_nomenclature in missing_in_residuals:
        new_row = {'Номенклатура': missing_nomenclature, 'Остаток': 0}
        residuals = residuals._append(new_row, ignore_index=True)

    # Обработка данных
    for index, row in residuals.iterrows():

        nomenclature = row["Номенклатура"]
        stock = row["Остаток"]

        sales_data_filtered_shop = sales[sales["Магазин"] == selected_store]


        if nomenclature in all_beer and end_date_beer != "None":
            sales_data_filtered_date = sales_data_filtered_shop[(sales_data_filtered_shop["Дата и время"] >= start_date_order) & (sales_data_filtered_shop["Дата и время"] <= end_date_beer)]
        elif nomenclature in kaspi_snacks and end_date_kaspi != "None":
            sales_data_filtered_date = sales_data_filtered_shop[(sales_data_filtered_shop["Дата и время"] >= start_date_order) & (sales_data_filtered_shop["Дата и время"] <= end_date_kaspi)]
        elif nomenclature in merka_snacks and end_date_merka != "None":
            sales_data_filtered_date = sales_data_filtered_shop[(sales_data_filtered_shop["Дата и время"] >= start_date_order) & (sales_data_filtered_shop["Дата и время"] <= end_date_merka)]
        elif nomenclature in banki and end_date_other != "None":
            sales_data_filtered_date = sales_data_filtered_shop[(sales_data_filtered_shop["Дата и время"] >= start_date_order) & (sales_data_filtered_shop["Дата и время"] <= end_date_other)]
        else:
            sales_data_filtered_date = sales_data_filtered_shop[(sales_data_filtered_shop["Дата и время"] >= start_date_str) & (sales_data_filtered_shop["Дата и время"] <= end_date_str)]


        matching_rows = sales_data_filtered_date[sales_data_filtered_date["Номенклатура"] == nomenclature]

        # print(nomenclature)
        # print(matching_rows)
        # print("\n\n\n")

        # Обработка данных для пива
        if not matching_rows.empty and matching_rows["Категория товара"].str.contains("Амбирлэнд").any() | matching_rows["Номенклатура"].str.contains("Квас").any():
            total_quantity = matching_rows["Количество товара"].sum()

            beer_result = handle_beer(nomenclature, stock, total_quantity)
            beer_results = beer_results._append(beer_result, ignore_index=True)
            # Замена именований
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Жигулёвское", "Жигулёвское Ф")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Светлое НФ", "Боровское светлое НФ")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Тёмное", "Боровское тёмное Ф")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Пилснер Ф", "Пилснер Ф")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Пилснер НФ", "Пилснер НФ")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Светлое Ф", "Бундес Ф")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Пшеничное", "Вайс канцлер НФ")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Вишнёвый крик", "Вайлд Черри")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Хеллес", "Боровское урожайное")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Крепкое", "Империал канцлер НФ")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Лёгкое", "Домашнее")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("IPA", "Хмельзилла ИПА НФ")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Грейпфрутовый эль", "Леди на велосипеде")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("APA", "Бирконг НФ АРА")
            beer_results["Номенклатура"] = beer_results["Номенклатура"].replace("Квас Воронеж", "Квас")


        # Обработка данных для закусок к пиву
        if not matching_rows.empty and matching_rows["Категория товара"].str.contains("Закуски к пиву").any() | matching_rows["Категория товара"].str.contains("Рыба").any():
            total_quantity = matching_rows["Количество товара"].sum()

            snacks_result = handle_snacks(nomenclature, stock, total_quantity)
            if nomenclature in kaspi_snacks:
                snacks_results_kaspi = snacks_results_kaspi._append(snacks_result, ignore_index=True)
            elif nomenclature in merka_snacks:
                snacks_results_merka = snacks_results_merka._append(snacks_result, ignore_index=True)
            else:
                print(f"{nomenclature} никуда не попало")

        if not matching_rows.empty and matching_rows["Категория товара"].str.contains("Прочее").any():
            total_quantity = matching_rows["Количество товара"].sum()

            other_result = handle_other(nomenclature, stock, total_quantity)
            other_results = other_results._append(other_result, ignore_index=True)

        if not any(sales_data_filtered_date["Номенклатура"] == nomenclature):
            if nomenclature in all_beer:
                beer_results = beer_results._append({"Номенклатура": nomenclature, "Остаток": stock, "Прогноз": 0, "Заказ кег": 0, "Остаток литров": 0}, ignore_index=True)
            elif nomenclature in merka_snacks:
                snacks_results_merka = snacks_results_merka._append({"Номенклатура": nomenclature, "Остаток": stock, "Прогноз": 0, "Прогнозируемый остаток": 0, "Заказ": 0}, ignore_index=True)
            elif nomenclature in kaspi_snacks:
                snacks_results_kaspi = snacks_results_kaspi._append({"Номенклатура": nomenclature, "Остаток": stock, "Прогноз": 0, "Прогнозируемый остаток": 0, "Заказ": 0}, ignore_index=True)
            elif nomenclature in banki:
                other_results = other_results._append({"Номенклатура": nomenclature, "Остаток": stock, "Прогноз": 0, "Прогнозируемый остаток": 0, "Заказ": 0}, ignore_index=True)



    #Вторая таблица для пива
    for index, row in beer_results.iterrows():
        nomenclature = row["Номенклатура"]
        forecast = row["Заказ кег"]

        if selected_store == "6_Люберцы_3-е Почтовое отделение74":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Домашнее", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Вайлд Черри", "Леди на велосипеде", "Квас"]
        elif selected_store == "16_Долгопрудный_Лихачевский68":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Бирконг НФ АРА", "Леди на велосипеде", "Вайлд Черри", "Квас", "Вице Канцлер б/а  0,45"]
        elif selected_store == "5_Балашиха_Фадеева3":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Домашнее", "Боровское урожайное", "Пилснер Ф", "Пилснер НФ", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Хмельзилла ИПА НФ", "Леди на велосипеде", "Вайлд Черри", "Квас"]
        elif selected_store == "7_Балашиха_Свердлова25":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Домашнее", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Хмельзилла ИПА НФ", "Леди на велосипеде", "Вайлд Черри", "Квас"]
        elif selected_store == "4_Электросталь_Ленина15":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Боровское урожайное", "Пилснер Ф", "Пилснер НФ", "Бундес Ф", "Вайс канцлер НФ", "Домашнее", "Империал канцлер НФ", "Бирконг НФ АРА", "Леди на велосипеде", "Вайлд Черри", "Квас"]
        elif selected_store == "3_Электросталь_Ялагина11":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Домашнее", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Леди на велосипеде", "Квас"]
        elif selected_store == "9_Балашиха_Советский6/17":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Домашнее", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Хмельзилла ИПА НФ", "Леди на велосипеде", "Вайлд Черри", "Квас"]
        elif selected_store == "14_Егорьевск_Советская191":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Домашнее", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Бирконг НФ АРА", "Леди на велосипеде", "Вайлд Черри", "Квас"]
        elif selected_store == "10_Коломна_Советская5":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Боровское урожайное", "Домашнее", "Пилснер Ф", "Пилснер НФ", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Хмельзилла ИПА НФ", "Леди на велосипеде", "Вайлд Черри", "Квас"]
        elif selected_store == "12_Фрязино_Мира8":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Боровское урожайное", "Домашнее", "Пилснер Ф", "Пилснер НФ", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Бирконг НФ АРА", "Леди на велосипеде", "Вайлд Черри", "Квас"]
        elif selected_store == "2_Электросталь_Победы1/2":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Домашнее", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Вайлд Черри", "Леди на велосипеде", "Квас"]
        elif selected_store == "1_Дрезна_Южная19а":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Домашнее", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Империал канцлер НФ", "Вайлд Черри", "Леди на велосипеде", "Квас"]
        elif selected_store == "18_Железнодорожный":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Вайлд Черри", "Квас"]
        elif selected_store == "17_Коломна_Кирова38":
            filred_snacks = ["Жигулёвское Ф", "Боровское светлое НФ", "Боровское тёмное Ф", "Пилснер Ф", "Бундес Ф", "Вайс канцлер НФ", "Вайлд Черри", "Бирконг НФ АРА", "Квас"]
        else:
            filred_snacks = []


        if nomenclature in filred_snacks:
            if nomenclature == "Пилснер НФ":
                beer_second_table = beer_second_table._append({beer_second_table.columns[0]: nomenclature, beer_second_table.columns[1]: f"{abs(forecast)}*50"}, ignore_index=True)
            elif nomenclature == "Пилснер Ф":
                beer_second_table = beer_second_table._append({beer_second_table.columns[0]: nomenclature, beer_second_table.columns[1]: f"{abs(forecast)}*50"}, ignore_index=True)
            elif nomenclature == "Жигулёвское Ф":
                beer_second_table = beer_second_table._append({beer_second_table.columns[0]: nomenclature, beer_second_table.columns[1]: f"{abs(forecast)}*50"}, ignore_index=True)
            elif nomenclature == "Квас":
                beer_second_table = beer_second_table._append({beer_second_table.columns[0]: nomenclature, beer_second_table.columns[1]: f"{abs(forecast)}*50"}, ignore_index=True)
            elif nomenclature == "Вайлд Черри":
                beer_second_table = beer_second_table._append({beer_second_table.columns[0]: nomenclature, beer_second_table.columns[1]: f"{abs(forecast)}*20"}, ignore_index=True)
            elif nomenclature == "Леди на велосипеде":
                beer_second_table = beer_second_table._append({beer_second_table.columns[0]: nomenclature, beer_second_table.columns[1]: f"{abs(forecast)}*20"}, ignore_index=True)
            elif nomenclature == "Вице Канцлер б/а  0,45":
                beer_second_table = beer_second_table._append({beer_second_table.columns[0]: nomenclature, beer_second_table.columns[1]: f"{abs(forecast)} уп."}, ignore_index=True)
            else:
                beer_second_table = beer_second_table._append({beer_second_table.columns[0]: nomenclature, beer_second_table.columns[1]: f"{abs(forecast)}*30"}, ignore_index=True)

    #Таблица с нашими названиями
    beer_second_table_our_name = beer_second_table.copy()

    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Жигулёвское Ф", "Жигулёвское")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Боровское светлое НФ", "Светлое НФ")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Боровское тёмное Ф", "Тёмное")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Пилснер Ф", "Пилснер Ф")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Пилснер НФ", "Пилснер НФ")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Бундес Ф", "Светлое Ф")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Вайс канцлер НФ", "Пшеничное")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Вайлд Черри", "Вишнёвый крик")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Боровское урожайное", "Хеллес")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Империал канцлер НФ", "Крепкое")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Домашнее", "Лёгкое")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Хмельзилла ИПА НФ", "IPA")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Леди на велосипеде", "Грейпфрутовый эль")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Бирконг НФ АРА", "APA")
    beer_second_table_our_name["Номенклатура(Воронеж)"] = beer_second_table_our_name["Номенклатура(Воронеж)"].replace("Квас", "Квас Воронеж")


    #Вторая таблица для закусок к пиву
    for index, row in snacks_results_merka.iterrows():
        nomenclature = row["Номенклатура"]
        forecast = row["Заказ"]
        forecasted_balance = row["Прогнозируемый остаток"]

        if nomenclature in piece_snacks:
            snacks_second_table = snacks_second_table._append({snacks_second_table.columns[0]: nomenclature, snacks_second_table.columns[1]: f"{int(forecast)} шт."}, ignore_index=True)
        else:
            snacks_second_table = snacks_second_table._append({snacks_second_table.columns[0]: nomenclature, snacks_second_table.columns[1]: f"{int(math.ceil(forecast))} кг."}, ignore_index=True)


    for index, row in snacks_results_kaspi.iterrows():
        nomenclature = row["Номенклатура"]
        forecast = row["Заказ"]
        forecasted_balance = row["Прогнозируемый остаток"]

        if nomenclature in piece_snacks:
            snacks_second_table_kaspi = snacks_second_table_kaspi._append({snacks_second_table_kaspi.columns[0]: nomenclature, snacks_second_table_kaspi.columns[1]: f"{int(forecast)} шт."}, ignore_index=True)
        else:
            snacks_second_table_kaspi = snacks_second_table_kaspi._append({snacks_second_table_kaspi.columns[0]: nomenclature, snacks_second_table_kaspi.columns[1]: f"{int(math.ceil(forecast))} кг."}, ignore_index=True)


    #Вторая таблица для прочего
    for index, row in other_results.iterrows():
        nomenclature = row["Номенклатура"]
        forecast = row["Заказ"]

        if nomenclature in banki:
            other_second_table = other_second_table._append({other_second_table.columns[0]: nomenclature, other_second_table.columns[1]: f"{int(forecast)} уп."}, ignore_index=True)
        else:
            other_second_table = other_second_table._append({other_second_table.columns[0]: nomenclature, other_second_table.columns[1]: f"{int(forecast)} шт."}, ignore_index=True)


    for static in static_nomenclature:
            new_row = {'Номенклатура': static, 'Заказ': "уп."}
            other_second_table = other_second_table._append(new_row, ignore_index=True)

    text_shop = text_shop._append(text_for_shop(), ignore_index=True)
    start_date_in_doc = start_date_in_doc._append({"":start_date_order.strftime('%d.%m.%Y %H:%M:%S')}, ignore_index=True)
    end_date_in_doc_kaspi = end_date_in_doc_kaspi._append({"":end_date_kaspi.strftime('%d.%m.%Y %H:%M:%S')}, ignore_index=True)
    end_date_in_doc_merka = end_date_in_doc_merka._append({"":end_date_merka.strftime('%d.%m.%Y %H:%M:%S')}, ignore_index=True)
    end_date_in_doc_beer = end_date_in_doc_beer._append({"":end_date_beer.strftime('%d.%m.%Y %H:%M:%S')}, ignore_index=True)
    end_date_in_doc_other = end_date_in_doc_other._append({"":end_date_other.strftime('%d.%m.%Y %H:%M:%S')}, ignore_index=True)

    save_file()

    # Запись таблиц в один Excel файл
    with pd.ExcelWriter(report_file_path) as writer:

        beer_results.to_excel(writer, sheet_name="Пиво", index=False)
        beer_second_table.to_excel(writer, sheet_name="Пиво", startrow=len(beer_results) + 3, index=False)
        beer_second_table_our_name.to_excel(writer, sheet_name="Пиво", startrow=len(beer_results) + 3, startcol=5, index=False)
        text_shop.to_excel(writer, sheet_name="Пиво", startcol=6, index=False)
        start_date_in_doc.to_excel(writer, sheet_name="Пиво", startcol=7, index=False)
        end_date_in_doc_beer.to_excel(writer, sheet_name="Пиво", startcol=8, index=False)

        snacks_results_merka.to_excel(writer, sheet_name="Мерка", index=False)
        snacks_second_table.to_excel(writer, sheet_name="Мерка", startrow=len(snacks_results_merka) + 3, index=False)
        text_shop.to_excel(writer, sheet_name="Мерка", startcol=6, index=False)
        start_date_in_doc.to_excel(writer, sheet_name="Мерка", startcol=7, index=False)
        end_date_in_doc_merka.to_excel(writer, sheet_name="Мерка", startcol=8, index=False)

        snacks_results_kaspi.to_excel(writer, sheet_name="Каспий", index=False)
        snacks_second_table_kaspi.to_excel(writer, sheet_name="Каспий", startrow=len(snacks_results_kaspi) + 3, index=False)
        text_shop.to_excel(writer, sheet_name="Каспий", startcol=6, index=False)
        start_date_in_doc.to_excel(writer, sheet_name="Каспий", startcol=7, index=False)
        end_date_in_doc_kaspi.to_excel(writer, sheet_name="Каспий", startcol=8, index=False)

        # snacks_results_sigma.to_excel(writer, sheet_name="Сиг", index=False)
        # snacks_second_table_sigma.to_excel(writer, sheet_name="Сиг", startrow=len(snacks_results_sigma) + 3, index=False)
        # text_shop.to_excel(writer, sheet_name="Сиг", startcol=6, index=False)
        # start_date_in_doc.to_excel(writer, sheet_name="Сиг", startcol=7, index=False)

        other_results.to_excel(writer, sheet_name="Прочее", index=False)
        other_second_table.to_excel(writer, sheet_name="Прочее", startrow=len(other_results) + 3, index=False)
        text_shop.to_excel(writer, sheet_name="Прочее", startcol=6, index=False)
        start_date_in_doc.to_excel(writer, sheet_name="Прочее", startcol=7, index=False)
        end_date_in_doc_other.to_excel(writer, sheet_name="Прочее", startcol=8, index=False)

        message_result = tk.Label(window, text="", foreground="blue")
        message_result = tk.Label(window, text=f"Файл {report_file_path.split("/")[-1]} загружен", foreground="blue")
        message_result.place(relx=0.5, rely=0.72, anchor="center")

    wb = load_workbook(report_file_path)

    ws_kaspi = wb["Каспий"]
    ws_merka = wb["Мерка"]

    fill_excel_kaspi(ws_kaspi)
    fill_excel_merka(ws_merka)

    wb.save(report_file_path)

    open_file (report_file_path)


window = tk.Tk()
window.title("Приложение")
window.geometry("500x350")
message_label_sales = tk.Label(window, text="Файл не выбран")
message_label_sales.place(x=65, y=40)
message_label_residuals = tk.Label(window, text="Файл не выбран")
message_label_residuals.place(x=340, y=40)
message_error = tk.Label(window, foreground="red")
message_error.place(relx=0.5, rely=0.23, anchor="center")

stores = tk.Label(window, text="Магазин")
stores.place(x=102, y=100)
store_combobox = ttk.Combobox(window, state="readonly")
store_combobox.place(x=105, y=120, width=300)


start_date_label = tk.Label(window, text="Начальная дата (в формате 29.03.2002)\nБудет поиск c 00:00:00")
start_date_label.place(x=10, y=155)
start_date = ttk.Entry(window)
start_date.place(x=55, y=200)

end_date_label = tk.Label(window, text="Конечная дата (в формате 29.03.2002)\nБудет поиск по 23:59:59")
end_date_label.place(x=280, y=155)
end_date = ttk.Entry(window)
end_date.place(x=320, y=200)


btn_sales = tk.Button(window, text="Загрузить файл с продажами", command=load_sales_file)
btn_residuals = tk.Button(window, text="Загрузить файл с остатками", command=load_residuals_file)
btn_sales.place(x=25, y=10)
btn_residuals.place(x=300, y=10)
btn_generate = tk.Button(window, text="Сгенерировать отчет", command=generate_report, background="#3498db", foreground="#FFFFFF", font="Arial 13 bold", borderwidth=0)
btn_generate.place(relx=0.5, rely=0.85, anchor="center", height=50, width=300)

window.mainloop()
